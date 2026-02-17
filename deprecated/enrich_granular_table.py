
import openpyxl
import argparse
import os
import shutil

def enrich_table(target_file, source_file):
    print(f"--- Enriching Granular Table ---")
    print(f"Target: {target_file}")
    print(f"Source: {source_file}")

    if not os.path.exists(target_file):
        print("Target file not found.")
        return
    if not os.path.exists(source_file):
        print("Source file not found.")
        return

    # 1. Load Source Data (RDC)
    print("Loading source data...")
    rdc_map = {}
    try:
        wb_src = openpyxl.load_workbook(source_file, read_only=True)
        ws_src = wb_src.active
        # Assuming Data starts at Row 1 or 2. Let's assume Row 1 if header is missing, or inspect.
        # Based on previous inspection, Row 1 seemed to be data (numeric ID 1).
        # We will iterate all rows.
        for row in ws_src.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            
            try:
                # Key: Column A (Index 0)
                rik_key = int(row[0]) 
                
                # Values: C(2), D(3), E(4), F(5)
                # Handle potential short rows
                c_val = row[2] if len(row) > 2 else None
                d_val = row[3] if len(row) > 3 else None
                e_val = row[4] if len(row) > 4 else None
                f_val = row[5] if len(row) > 5 else None
                
                rdc_map[rik_key] = {
                    'C': c_val, # Rishi
                    'D': d_val, # Chandas
                    'E': e_val, # Devata
                    'F': f_val  # Metadata/Other
                }
            except ValueError:
                # Header row or invalid key
                continue
        wb_src.close()
        print(f"Loaded {len(rdc_map)} records from source.")
        
    except Exception as e:
        print(f"Error reading source: {e}")
        return

    # 2. Update Target Data (Granular Table)
    print("Updating target table...")
    try:
        # Load workbook (not read_only so we can edit)
        wb_tgt = openpyxl.load_workbook(target_file)
        ws_tgt = wb_tgt.active
        
        # Identify Column Indices (0-based) based on inspect_headers output
        # Global_Rik_Num is Index 1
        # Rik_Rishi is Index 9
        # Rik_Devata is Index 10
        # Rik_Chandas is Index 11
        # Rik_Metadata is Index 12
        
        update_count = 0
        
        # Iterate rows, skipping headers (assuming 2 header rows based on previous knowledge, specifically row 2 has names)
        # We start from row 3 (1-based index 3)
        for row in ws_tgt.iter_rows(min_row=3):
            # Global_Rik_Num is at column 2 (1-based index) -> row[1]
            cell_rik_num = row[1]
            rik_num_val = cell_rik_num.value
            
            if rik_num_val is None:
                continue
                
            try:
                rik_num = int(rik_num_val)
            except ValueError:
                continue
            
            if rik_num in rdc_map:
                src_data = rdc_map[rik_num]
                
                # Update Columns
                # C (Rishi) -> Rik_Rishi (Index 9 -> Column 10)
                # D (Chandas) -> Rik_Chandas (Index 11 -> Column 12)
                # E (Devata) -> Rik_Devata (Index 10 -> Column 11)
                # F (Meta) -> Rik_Metadata (Index 12 -> Column 13)
                
                # Update Rik_Rishi
                if src_data['C']: row[9].value = src_data['C']
                
                # Update Rik_Devata (Source Col E)
                if src_data['E']: row[10].value = src_data['E']
                
                # Update Rik_Chandas (Source Col D)
                if src_data['D']: row[11].value = src_data['D']
                
                # Update Rik_Metadata (Source Col F)
                if src_data['F']: row[12].value = src_data['F']
                
                update_count += 1
                
        # Save
        backup_file = target_file.replace(".xlsx", ".bak.xlsx")
        shutil.copyfile(target_file, backup_file)
        print(f"Backup saved to {backup_file}")
        
        wb_tgt.save(target_file)
        print(f"Successfully updated {update_count} rows in {target_file}")
        wb_tgt.close()

    except Exception as e:
        print(f"Error updating target: {e}")

if __name__ == "__main__":
    base_dir = r"c:\Users\sekha\OneDrive\Documents\GitHub\jaimineeyasamavedam"
    target = os.path.join(base_dir, r"data\output\JSV_Samam_Granular_Table.xlsx")
    source = os.path.join(base_dir, r"data\input\Samved-RDC.xlsx")
    enrich_table(target, source)
