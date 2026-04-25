"""
Generate a Rik-level table listing every unique Rik.
This table contains one row per Rik, avoiding n:1 Samam-to-Rik duplication issues.

Usage:
    python src/generate_rik_table.py [input_json] [-o output_csv]

Examples:
    # Use defaults:
    python src/generate_rik_table.py

    # Specify input file:
    python src/generate_rik_table.py data/output/Samhita_Devanagari_Unicode_out.json

    # Specify both input and output:
    python src/generate_rik_table.py data/output/Samhita_Devanagari_Unicode_out.json -o data/output/My_Rik_Table.csv
"""
import argparse
import json
import csv
import os
import sys
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

from utils import get_generated_metadata

# Try to import openpyxl for Excel handling
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

DEFAULT_INPUT = r'data\output\Samhita_corrected_out.json'
DEFAULT_OUTPUT_CSV = r'data\output\JSV_Rik_Table.csv'
DEFAULT_RECON_EXCEL = r'data\output\Rik Reconciliation table (JSV-KSV).xlsx'
DEFAULT_VARGEEKARAN_JSON = r'data\output\Vargeekaran.json'

def replace_accents_unicode(text):
    """
    Replaces ASCII parenthesis accent markers with actual Unicode accents.
    (1): Swarita (U+0951)
    (2): Anudatta (U+1CD2)
    (3): Kampa (U+1CF8)
    (4): Trikampa (U+1CF9)
    """
    if not text:
        return text
    replacements = [
        ('(1)', '\u0951'),
        ('(2)', '\u1CD2'),
        ('(3)', '\u1CF8'),
        ('(4)', '\u1CF9'),
    ]
    for marker, unicode_val in replacements:
        text = text.replace(marker, unicode_val)
    return text

def load_reconciliation_data(excel_path):
    """Load Rishi, Chandas, Devata mapping from Excel."""
    mapping = {} # Global_Rik_Num -> (Rishi, Chandas, Devata)
    if not excel_path or not os.path.exists(excel_path):
        print(f"[WARNING] Reconciliation Excel not found at: {excel_path}")
        return mapping
    
    if not HAS_OPENPYXL:
        print("[ERROR] 'openpyxl' is required to read Excel files. Please install it.")
        return mapping

    print(f"Loading reconciliation data from {excel_path}...")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # We expect: 
        # Column A: Global_Rik_Num (Index 0)
        # Column I: Rishi (Index 8)
        # Column J: Chandas (Index 9)
        # Column K: Devata (Index 10)
        
        row_count = 0
        # Data starts from row 3 (Row 1: Metadata, Row 2: Headers)
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 11:
                continue
            
            try:
                # Row index 0 is Column A
                if row[0] is not None:
                    # Handle both integer and string (in case Excel formatted it weirdly)
                    rik_num = int(row[0])
                    # Row index 5 is F (Rik_Metadata)
                    # Row index 8 is I (Rishi), 9 is J (Chandas), 10 is K (Devata)
                    rik_metadata_val = str(row[5]).strip() if row[5] is not None else ""
                    rishi = str(row[8]).strip() if row[8] is not None else ""
                    chandas = str(row[9]).strip() if row[9] is not None else ""
                    devata = str(row[10]).strip() if row[10] is not None else ""
                    
                    mapping[rik_num] = {
                        "rishi": rishi,
                        "chandas": chandas,
                        "devata": devata,
                        "rik_metadata": rik_metadata_val
                    }
                    row_count += 1
            except (ValueError, TypeError):
                continue
                
        print(f"Successfully loaded {row_count} classification entries from Excel.")
    except Exception as e:
        print(f"[ERROR] Failed to read Excel: {e}")
        
    return mapping

def main(mode='samhita', input_file=None, output_csv=None, recon_excel=None, v_json=None, no_enrich=False):
    # 0. Load Configuration
    from utils import load_pipeline_config
    pipeline_cfg = load_pipeline_config()
    table_cfg = pipeline_cfg.get('generate_rik_table', {})
    
    # Use the provided mode, or the one from config, or default to 'samhita'
    active_mode = mode or table_cfg.get('default_type') or 'samhita'
    mode_cfg = table_cfg.get(active_mode, {})

    # Priority: Function Args > Mode Config > Root Config > Default Constants
    input_file = input_file or mode_cfg.get('input') or table_cfg.get('input') or DEFAULT_INPUT
    output_csv = output_csv or mode_cfg.get('output_csv') or table_cfg.get('output_csv') or DEFAULT_OUTPUT_CSV
    recon_excel = recon_excel or mode_cfg.get('recon_excel') or table_cfg.get('recon_excel') or DEFAULT_RECON_EXCEL
    v_json = v_json or mode_cfg.get('v_json') or table_cfg.get('v_json') or DEFAULT_VARGEEKARAN_JSON

    # Get metadata
    metadata = get_generated_metadata()
    JSV_VERSION = metadata['version']
    GENERATED_AT = metadata['generated_at']

    print(f"Generating Rik Table (v{JSV_VERSION})...")
    print(f"Input JSON : {input_file}")
    print(f"Output CSV : {output_csv}")
    if not no_enrich:
        print(f"Recon Excel: {recon_excel}")
    else:
        print(f"Recon Excel: [SKIPPED - No Enrich Mode]")
    print(f"V-JSON Out : {v_json}")

    # Load the JSON
    if not os.path.exists(input_file):
        print(f"Error: Input file '{input_file}' not found.")
        return

    with open(input_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Use cascading metadata if present in JSON, otherwise fallback
    json_meta = data.get('meta', {})
    if json_meta.get('version'):
        JSV_VERSION = json_meta['version']
        print(f"[INFO] Using cascading Version {JSV_VERSION}")
    
    # Always use fresh timestamp for the export
    GENERATED_at = GENERATED_AT

    # Load Excel Reconciliation Data
    recon_data = {}
    if not no_enrich:
        recon_data = load_reconciliation_data(recon_excel)

    # CSV rows
    rows = []
    
    # Vargeekaran JSON data
    vargeekaran_data = []

    # Get the supersection container
    supersection_container = data.get('supersection', {})

    # Sort supersection keys numerically
    ss_keys = sorted(supersection_container.keys(), key=lambda x: int(x.split('_')[1]) if '_' in x else 0)

    excel_pointer = 0
    unique_riks = {}
    prev_rik_id_global_context = None

    for ss_key in ss_keys:
        ss_data = supersection_container[ss_key]
        ss_title = ss_data.get('supersection_title', ss_key)
        
        # Get sections
        sections = ss_data.get('sections', {})
        sec_keys = sorted(sections.keys(), key=lambda x: int(x.split('_')[1]) if '_' in x else 0)
        
        for sec_key in sec_keys:
            sec_data = sections[sec_key]
            sec_title = sec_data.get('section_title', sec_key)
            display_rik_counter = 0
            
            # Get subsections (Arsheyams/Riks)
            subsections = sec_data.get('subsections', {})
            sub_keys = sorted(subsections.keys(), key=lambda x: int(x.split('_')[1]) if '_' in x else 0)
            
            for sub_key in sub_keys:
                sub_data = subsections[sub_key]
                
                # Get Rik info
                try:
                    base_rik_id = int(sub_data.get('rik_id', 0))
                except ValueError:
                    base_rik_id = 0
                    
                rik_metadata = sub_data.get('rik_metadata', '')
                rik_text_full = sub_data.get('rik_text', '')

                # 1. Handle Empty Rik Text (Explicit null or inherited null)
                if not rik_text_full.strip():
                    excel_pointer += 1
                    classification = recon_data.get(excel_pointer, {})
                    
                    # For "null" Rik text, we do not increment display counter or show a number
                    rik_metadata = classification.get('rik_metadata', "")
                    sub_data['rik_metadata'] = rik_metadata
                    
                    sub_data['rik_classifications'] = [{
                        "Global_Rik_Num": excel_pointer,
                        "Rik_ID": "null",
                        "Rishi": classification.get('rishi', ""),
                        "Chandas": classification.get('chandas', ""),
                        "Devata": classification.get('devata', ""),
                        "Rik_Metadata": rik_metadata
                    }]

                    rows.append({
                        'Global_Rik_Num': excel_pointer,
                        'Patha_Name': ss_title,
                        'Khanda': sec_title,
                        'Rik_ID': "null",
                        'Rishi': classification.get('rishi', ""),
                        'Chandas': classification.get('chandas', ""),
                        'Devata': classification.get('devata', ""),
                        'Rik_Text': "null",
                        'Rik_Metadata': rik_metadata
                    })
                    continue

                # 2. Process Non-Empty Rik Text
                import re
                # Robust split: find all markers (॥ N ॥) and treat the text preceding them as that Rik.
                # This correctly handles cases where multiple Riks are on the same line.
                parts = re.split(r'(॥\s*[०-९\d]+\s*॥)', rik_text_full)
                
                # pairs of (text_before, marker)
                for i in range(0, len(parts) - 1, 2):
                    text_seg = parts[i].strip()
                    marker_seg = parts[i+1].strip()
                    
                    # If this is the last marker, append any trailing text from the very last element of re.split
                    if i + 2 == len(parts) - 1:
                        trailing = parts[i+2].strip()
                        if trailing:
                             marker_seg += " " + trailing

                    # Identify the Rik ID from the marker
                    match = re.search(r'([०-९\d]+)', marker_seg)
                    if match:
                        num_str = match.group(1)
                        # convert devanagari to int
                        devanagari_digits = '०१२३४५६७८९'
                        for j, char in enumerate(devanagari_digits):
                            num_str = num_str.replace(char, str(j))
                        try:
                            line_rik_id = int(num_str)
                        except ValueError:
                            line_rik_id = base_rik_id
                    else:
                        line_rik_id = "null"

                    combined_text = (text_seg + " " + marker_seg).strip()
                    current_rik_key = (ss_key, sec_key, line_rik_id)
                    
                    # Rule: Only increment excel_pointer for NEW unique Riks
                    # Repeated Riks in different Samams will share the same Global_Rik_Num
                    if current_rik_key not in unique_riks:
                        excel_pointer += 1
                        display_rik_counter += 1
                        
                        classification = recon_data.get(excel_pointer, {})
                        
                        # Rule: If JSON has explicit null metadata but non-null Rik text, do NOT override with Excel
                        if rik_metadata == "" and combined_text.strip() != "":
                             # Keep it null
                             pass
                        else:
                             # Update JSON metadata strictly from Excel if present
                             if classification.get('rik_metadata'):
                                 rik_metadata = classification.get('rik_metadata', "")
                                 sub_data['rik_metadata'] = rik_metadata
                            
                        # Add row for this unique Rik to the CSV (unique list)
                        # Replace ASCII markers with Unicode accents
                        clean_text = replace_accents_unicode(combined_text)
                        row_entry = {
                            'Global_Rik_Num': excel_pointer,
                            'Patha_Name': ss_title,
                            'Khanda': sec_title,
                            'Rik_ID': display_rik_counter,
                            'Rishi': classification.get('rishi', ""),
                            'Chandas': classification.get('chandas', ""),
                            'Devata': classification.get('devata', ""),
                            'Rik_Text': clean_text,
                            'Rik_Metadata': rik_metadata
                        }
                        unique_riks[current_rik_key] = {
                            'Global_Rik_Num': excel_pointer,
                            'Rik_ID': display_rik_counter,
                            'classification': classification,
                            'rik_metadata_override': rik_metadata
                        }
                        rows.append(row_entry)
                    
                    # Use the stored shared info for this Rik (even if repeated)
                    rik_info = unique_riks[current_rik_key]
                    final_global_num = rik_info['Global_Rik_Num']
                    final_rik_id = rik_info['Rik_ID']
                    final_metadata = rik_info['rik_metadata_override']
                    final_class = rik_info['classification']

                    # Always inject into sub_data for Vargeekaran JSON
                    if 'rik_classifications' not in sub_data:
                        sub_data['rik_classifications'] = []
                    
                    sub_data['rik_classifications'].append({
                        "Global_Rik_Num": final_global_num,
                        "Rik_ID": final_rik_id,
                        "Rishi": final_class.get('rishi', ""),
                        "Chandas": final_class.get('chandas', ""),
                        "Devata": final_class.get('devata', ""),
                        "Rik_Metadata": final_metadata
                    })

    # Write CSV with UTF-8 BOM for Excel compatibility
    with open(output_csv, 'w', encoding='utf-8-sig', newline='') as f:
        # Line 1: <Filename> <Version> <Timestamp>
        filename = os.path.basename(output_csv)
        f.write(f"{filename} {JSV_VERSION} {GENERATED_AT}\n")

        fieldnames = [
            'Global_Rik_Num', 'Patha_Name', 'Khanda', 'Rik_ID', 'Rik_Text', 'Rik_Metadata'
        ]
        
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(rows)

    print(f"CSV saved to: {output_csv}")
    
    # Write Excel if pandas is available
    if HAS_PANDAS:
        from openpyxl.styles import Font
        output_xlsx = output_csv.rsplit('.', 1)[0] + '.xlsx'
        try:
            # Create DataFrame
            df = pd.DataFrame(rows)
            # Reorder columns to match CSV preference
            cols = ['Global_Rik_Num', 'Patha_Name', 'Khanda', 'Rik_ID', 'Rishi', 'Chandas', 'Devata', 'Rik_Text', 'Rik_Metadata']
            # Only use columns that actually exist in the rows
            cols = [c for c in cols if c in df.columns]
            df = df[cols]
            
            # Save to Excel
            with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
                # 1. Create and save Metadata sheet
                meta_rows = [
                    ["Project", "Jaimineeya Samavedam"],
                    ["Filename", os.path.basename(output_xlsx)],
                    ["Version", JSV_VERSION],
                    ["Generated At", GENERATED_AT]
                ]
                meta_df = pd.DataFrame(meta_rows, columns=["Property", "Value"])
                meta_df.to_excel(writer, index=False, sheet_name='Metadata')
                
                # Formatting Metadata sheet
                meta_ws = writer.sheets['Metadata']
                standard_font = Font(name='Adishila', size=11)
                bold_font = Font(name='Adishila', size=11, bold=True)
                
                for row in meta_ws.iter_rows():
                    for cell in row:
                        cell.font = standard_font
                
                # Bold headers for Metadata
                for cell in meta_ws[1]:
                    cell.font = bold_font
                    
                meta_ws.column_dimensions['A'].width = 15
                meta_ws.column_dimensions['B'].width = 40

                # 2. Save main Rik Table
                df.to_excel(writer, index=False, sheet_name='Rik Table')
                
                # Formatting Rik Table sheet
                worksheet = writer.sheets['Rik Table']
                
                # Apply Adishila to all cells
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.font = standard_font
                
                # Bold headers for Rik Table
                for cell in worksheet[1]:
                    cell.font = bold_font
                
                # Auto-adjust column widths for Rik Table
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter # Get the column name
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column].width = min(adjusted_width, 100) # Cap at 100

            print(f"Excel table saved to: {output_xlsx}")
        except Exception as e:
            print(f"[ERROR] Failed to save Excel: {e}")
    else:
        print("[WARNING] 'pandas' not found. Skipping Excel export.")
    
    # Write Enhanced Vargeekaran JSON (Full Samhita structure with injected info)
    with open(v_json, 'w', encoding='utf-8') as f:
        # Update meta if needed or keep original
        if 'meta' not in data:
            data['meta'] = {}
        data['meta']['description'] = "Enhanced Samhita with Rishi, Devata, Chandas classification"
        data['meta']['generated_at'] = GENERATED_AT
        data['meta']['version'] = JSV_VERSION

        json.dump(data, f, indent=4, ensure_ascii=False)
    print(f"Vargeekaran JSON saved to: {v_json}")
    
    print(f"Total Unique Riks: {len(rows)}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate a Rik-level table listing every unique Rik."
    )
    parser.add_argument("--type", choices=["samhita", "aaranam"],
                        help="Mode of operation: 'samhita' (default) or 'aaranam'")
    parser.add_argument("input", nargs="?",
                        help="Override Input JSON file")
    parser.add_argument("-o", "--output",
                        help="Override Output CSV file")
    parser.add_argument("-e", "--excel",
                        help="Override Reconciliation Excel file")
    parser.add_argument("-j", "--json_out",
                        help="Override Output Vargeekaran JSON file")

    parser.add_argument("--no-enrich", action="store_true",
                        help="Skip enrichment from external Excel reconciliation table")

    args = parser.parse_args()
    main(mode=args.type, input_file=args.input, output_csv=args.output, recon_excel=args.excel, v_json=args.json_out, no_enrich=args.no_enrich)
